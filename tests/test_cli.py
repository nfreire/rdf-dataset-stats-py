import logging
from pathlib import Path

from openpyxl import load_workbook

from rdf_dataset_stats import cli
from rdf_dataset_stats.model import DatasetStats


def make_stats() -> DatasetStats:
    stats = DatasetStats()
    stats.increment_class("http://example.org/Class", amount=2)
    return stats


def test_cli_successful_positional_invocation(monkeypatch, tmp_path, capsys) -> None:
    input_path = tmp_path / "dump"
    input_path.mkdir()
    output_path = tmp_path / "stats.xlsx"

    def fake_collect_dump_stats(path: Path, **kwargs):
        assert path == input_path
        return make_stats(), 7

    monkeypatch.setattr(cli, "collect_dump_stats", fake_collect_dump_stats)

    exit_code = cli.main([str(input_path), str(output_path)])

    assert exit_code == 0
    assert output_path.exists()
    summary = capsys.readouterr().out
    assert "Processed records: 7" in summary
    assert "Classes found: 1" in summary
    assert f"Output written to: {output_path}" in summary
    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Class"]


def test_cli_successful_long_option_invocation(monkeypatch, tmp_path, capsys) -> None:
    input_path = tmp_path / "dump"
    input_path.mkdir()
    output_path = tmp_path / "stats.xlsx"
    monkeypatch.setattr(
        cli, "collect_dump_stats", lambda path, **kwargs: (make_stats(), 3)
    )

    exit_code = cli.main(
        ["--input", str(input_path), "--output", str(output_path)]
    )

    assert exit_code == 0
    assert output_path.exists()
    assert "Processed records: 3" in capsys.readouterr().out


def test_cli_missing_arguments_returns_nonzero(capsys) -> None:
    try:
        cli.main([])
    except SystemExit as exc:
        exit_code = exc.code
    else:
        exit_code = 0

    assert exit_code != 0
    assert "input and output paths are required" in capsys.readouterr().err


def test_cli_rejects_invalid_input_directory(tmp_path, capsys) -> None:
    missing_input = tmp_path / "missing"
    output_path = tmp_path / "stats.xlsx"

    exit_code = cli.main([str(missing_input), str(output_path)])

    assert exit_code == 1
    assert not output_path.exists()
    assert "input path is not a directory" in capsys.readouterr().err


def test_cli_reports_unrecoverable_error_without_traceback(
    monkeypatch, tmp_path, capsys
) -> None:
    input_path = tmp_path / "dump"
    input_path.mkdir()
    output_path = tmp_path / "stats.xlsx"

    def fake_collect_dump_stats(path: Path, **kwargs):
        raise OSError("cannot read dump")

    monkeypatch.setattr(cli, "collect_dump_stats", fake_collect_dump_stats)

    exit_code = cli.main([str(input_path), str(output_path)])

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "Error: cannot read dump" in captured.err
    assert "Traceback" not in captured.err


def test_cli_prints_subdataset_and_progress_updates(
    monkeypatch, tmp_path, capsys
) -> None:
    input_path = tmp_path / "dump"
    input_path.mkdir()
    output_path = tmp_path / "stats.xlsx"

    def fake_collect_dump_stats(path: Path, **kwargs):
        kwargs["on_subdataset"]("00719")
        kwargs["on_progress"](150, 2.5, 400_000)
        return make_stats(), 150

    monkeypatch.setattr(cli, "collect_dump_stats", fake_collect_dump_stats)

    exit_code = cli.main([str(input_path), str(output_path)])

    output = capsys.readouterr().out
    assert exit_code == 0
    assert "Processing subdataset: 00719" in output
    assert "Records processed so far: 150" in output
    assert "Average records per second: 2.50" in output
    assert "Estimated time for 1000000 records: 111:06:40" in output


def test_cli_suppresses_noisy_rdflib_term_logging(
    monkeypatch, tmp_path
) -> None:
    input_path = tmp_path / "dump"
    input_path.mkdir()
    output_path = tmp_path / "stats.xlsx"
    logger = logging.getLogger("rdflib.term")
    previous_level = logger.level
    logger.setLevel(logging.NOTSET)
    monkeypatch.setattr(
        cli, "collect_dump_stats", lambda path, **kwargs: (make_stats(), 1)
    )

    try:
        exit_code = cli.main([str(input_path), str(output_path)])
    finally:
        logger.setLevel(previous_level)

    assert exit_code == 0
    assert logger.level == logging.CRITICAL


def test_cli_writes_intermediate_and_final_output(
    monkeypatch, tmp_path
) -> None:
    input_path = tmp_path / "dump"
    input_path.mkdir()
    output_path = tmp_path / "stats.xlsx"
    write_calls: list[Path] = []

    def fake_collect_dump_stats(path: Path, **kwargs):
        kwargs["on_intermediate_result"](make_stats(), 1_000_000)
        return make_stats(), 1_000_001

    def fake_write_excel(stats: DatasetStats, path: Path) -> None:
        write_calls.append(path)
        path.write_text("xlsx placeholder")

    monkeypatch.setattr(cli, "collect_dump_stats", fake_collect_dump_stats)
    monkeypatch.setattr(cli, "write_excel", fake_write_excel)

    exit_code = cli.main([str(input_path), str(output_path)])

    assert exit_code == 0
    assert write_calls == [output_path, output_path]
    assert output_path.exists()
