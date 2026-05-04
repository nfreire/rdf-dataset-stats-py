from openpyxl import load_workbook

from rdf_dataset_stats.excel import write_excel
from rdf_dataset_stats.model import DatasetStats


def test_write_excel_includes_class_metadata_and_counts(tmp_path) -> None:
    stats = DatasetStats()
    stats.increment_class("http://example.org/Class", amount=3)
    stats.increment_literal_property(
        "http://example.org/Class", "http://example.org/title", amount=2
    )
    stats.increment_resource_property(
        "http://example.org/Class",
        "http://example.org/related",
        "http://example.org/ObjectClass",
        amount=4,
    )
    output_path = tmp_path / "stats.xlsx"

    write_excel(stats, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["Class"]
    assert sheet["A1"].value == "Class URI:"
    assert sheet["B1"].value == "http://example.org/Class"
    assert sheet["A2"].value == "Class count:"
    assert sheet["B2"].value == 3
    assert [cell.value for cell in sheet[4]] == [
        "Property URI",
        "Literal count",
        "Resource object class URI",
        "Resource count",
    ]
    assert [cell.value for cell in sheet[5]] == [
        "http://example.org/related",
        0,
        None,
        None,
    ]
    assert [cell.value for cell in sheet[6]] == [
        "http://example.org/related",
        None,
        "http://example.org/ObjectClass",
        4,
    ]
    assert [cell.value for cell in sheet[7]] == [
        "http://example.org/title",
        2,
        None,
        None,
    ]


def test_write_excel_orders_sheets_properties_and_resource_classes(tmp_path) -> None:
    stats = DatasetStats()
    stats.increment_class("http://example.org/Z")
    stats.increment_class("http://example.org/A")
    stats.increment_literal_property("http://example.org/Z", "http://example.org/b")
    stats.increment_literal_property("http://example.org/Z", "http://example.org/a")
    stats.increment_resource_property(
        "http://example.org/Z", "http://example.org/b", "http://example.org/ObjectZ"
    )
    stats.increment_resource_property(
        "http://example.org/Z", "http://example.org/b", "http://example.org/ObjectA"
    )
    output_path = tmp_path / "ordered.xlsx"

    write_excel(stats, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["A", "Z"]
    sheet = workbook["Z"]
    assert sheet["A5"].value == "http://example.org/a"
    assert sheet["A6"].value == "http://example.org/b"
    assert sheet["C7"].value == "http://example.org/ObjectA"
    assert sheet["C8"].value == "http://example.org/ObjectZ"


def test_write_excel_uses_unique_safe_sheet_names(tmp_path) -> None:
    stats = DatasetStats()
    class_uris = [
        "http://example.org/path/Name:With/Invalid?Characters*And[Long]SuffixOne",
        "http://another.example/path/Name:With/Invalid?Characters*And[Long]SuffixTwo",
        "http://example.org/Name:With/Invalid?Characters*And[Long]SuffixOne",
    ]
    for class_uri in class_uris:
        stats.increment_class(class_uri)
    output_path = tmp_path / "safe-names.xlsx"

    write_excel(stats, output_path)

    workbook = load_workbook(output_path)
    assert len(workbook.sheetnames) == 3
    assert len(set(workbook.sheetnames)) == 3
    for sheet_name in workbook.sheetnames:
        assert len(sheet_name) <= 31
        assert not any(char in sheet_name for char in ':\\/?*[]')

    ordered_class_uris = sorted(class_uris)
    assert [workbook[name]["B1"].value for name in workbook.sheetnames] == ordered_class_uris
